import openpyxl
import time
import os
import numpy as np

wb = openpyxl.load_workbook('excel.xlsx')
sh1 = wb['vlookup2']
list_with_values=[]
for cell in sh1[1]:
    list_with_values.append(cell.value)
print(list_with_values)
sheet = wb['vlookup2']
row = sheet.max_row
lis = []
for i in range(2, row+1):
    lis.append(sheet.cell(row = i, column = 2).value)
print(lis)
wb1 = openpyxl.workbook.Workbook()
wb1.remove(wb1['Sheet'])

for i in np.unique(lis):
    print(i)
    locals()['ws_'+str(i)] = wb1.create_sheet(i)
    locals()['ws_'+str(i)].title = i
    locals()['sh_'+str(i)] = wb1[i]
    locals()['sh_'+str(i)].append(list_with_values)
    for row in sh1.iter_rows():
        if row[1].value == i:
            locals()['sh_'+str(i)].append((cell.value for cell in row))

wb1.save(filename = 'sample.xlsx')
            